from app import create_app
from flask import current_app
from flask.cli import with_appcontext
import click
from app.db import db
from app.models import User, Membro
import os

app = create_app()

@app.cli.command('create-admin')
@click.option('--name', prompt=True)
@click.option('--email', prompt=True)
@click.option('--password', prompt=True, hide_input=True)
@with_appcontext
def create_admin(name, email, password):
	user = User.query.filter_by(email=email.lower()).first()
	if user:
		click.echo('Usuário já existe')
		return
	user = User(name=name, email=email.lower(), role='admin')
	user.set_password(password)
	db.session.add(user)
	db.session.commit()
	click.echo(f'Admin criado: {email}')

@app.cli.command('import-membros')
@click.argument('path')
@with_appcontext
def import_membros(path):
	path = os.path.abspath(path)
	if not os.path.exists(path):
		click.echo(f'Arquivo não encontrado: {path}')
		return
	rows = []
	ext = os.path.splitext(path)[1].lower()
	if ext == '.xlsx':
		from openpyxl import load_workbook
		wb = load_workbook(path)
		sheet = wb.active
		headers = [str(c.value or '').strip() for c in next(sheet.iter_rows(min_row=1, max_row=1))[0:]]
		rows = list(sheet.iter_rows(min_row=2, values_only=True))
	elif ext == '.xls':
		import xlrd
		wb = xlrd.open_workbook(path)
		sheet = wb.sheet_by_index(0)
		headers = [str(sheet.cell_value(0, col) or '').strip() for col in range(sheet.ncols)]
		for r in range(1, sheet.nrows):
			rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
	else:
		click.echo('Formato não suportado. Use .xls ou .xlsx')
		return

	idx = { h:i for i,h in enumerate(headers) }
	ins = 0
	for r in rows:
		def v(h):
			i = idx.get(h)
			val = (str(r[i]).strip() if i is not None and r[i] is not None else None)
			return val
		m = Membro(
			nome = v('Membro') or v('Nome'),
			sexo = v('Sexo'),
			concurso = v('Concurso'),
			cargo_efetivo = v('Cargo efetivo'),
			titularidade = v('Titularidade'),
			email_pessoal = v('eMail pessoal') or v('Email') or v('E-mail'),
			cargo_especial = v('Cargo Especial'),
			telefone_unidade = v('Telefone Unidade'),
			telefone_celular = v('Telefone celular'),
			unidade_lotacao = v('Unidade Lotação'),
			comarca_lotacao = v('Comarca Lotação'),
			time_extraprofissionais = v('Time de futebol e outros grupos extraprofissionais'),
			quantidade_filhos = (int(float(v('Quantidade de filhos'))) if (v('Quantidade de filhos') or '').strip() not in ('', None) else None),
			nomes_filhos = v('Nome dos filhos'),
			estado_origem = (v('Estado de origem') or '')[:2] or None,
			academico = v('Acadêmico') or v('Academico'),
			pretensao_carreira = v('Pretensão de movimentação na carreira') or v('Pretensao de movimentacao na carreira'),
			carreira_anterior = v('Carreira anterior'),
			lideranca = v('Liderança') or v('Lideranca'),
			grupos_identitarios = v('Grupos identitários') or v('Grupos identitarios'),
		)
		db.session.add(m)
		ins += 1
	db.session.commit()
	click.echo(f'Importados: {ins}')

@app.cli.command('seed-demo')
@with_appcontext
def seed_demo():
	# cria alguns membros de exemplo
	if Membro.query.count() > 0:
		click.echo('Já existem membros, não será duplicado.')
		return
	m1 = Membro(nome='WILSON PENIN COUTO', sexo='Masculino', concurso='2001', cargo_efetivo='Promotor', email_pessoal='wilson@example.com', comarca_lotacao='BELO HORIZONTE')
	m2 = Membro(nome='WESLEY LEITE VAZ', sexo='Masculino', concurso='2002', cargo_efetivo='Promotor', email_pessoal='wesley@example.com', comarca_lotacao='BELO HORIZONTE')
	m3 = Membro(nome='VANESSA CAMPOLINA REBELLO HORTA', sexo='Feminino', concurso='2003', cargo_efetivo='Promotora', email_pessoal='vanessa@example.com', comarca_lotacao='CONTAGEM')
	db.session.add_all([m1,m2,m3])
	db.session.commit()
	# relacionamentos (amigos no MP)
	m1.amigos.append(m2)
	m1.amigos.append(m3)
	db.session.commit()
	click.echo('Seed de membros concluído com 3 registros e relacionamentos.')

if __name__ == '__main__':
	app.run(host='0.0.0.0', port=8000, debug=True) 